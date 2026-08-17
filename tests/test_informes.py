# -*- coding: utf-8 -*-
"""
================================================================================
 GUARDA DE INFORMES EXPORTABLES  --  el archivo trae lo que el agregado calculo
================================================================================

Por que existe
--------------
nucleo/herramientas/informes.py convierte el resultado YA CALCULADO de una
herramienta 'agregado' (nucleo/herramientas/agregado.py) en un archivo
(Excel o PDF). No debe agregar NINGUN numero que el agregado no haya
calculado -- si lo hiciera, estaria violando PRD 12.5 ('el modelo compone,
el codigo calcula') por la puerta de atras: el archivo tendria un dato que
nadie calculo con el metodo verificado.

La parte de PDF usa pypdf SOLO para leer el archivo de vuelta y verificar su
texto -- no es una dependencia del motor (que solo ESCRIBE PDFs, con
reportlab), instalarla aparte si falta (ver requirements.txt).

Uso
---
    py -3.13 tests/test_informes.py
================================================================================
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from nucleo.herramientas import informes   # noqa: E402

fallos: list[str] = []


def comprobar(condicion: bool, que: str) -> None:
    print(f"  {'[ok]  ' if condicion else '[FALLA]'} {que}")
    if not condicion:
        fallos.append(que)


def _leer(bytes_xlsx: bytes) -> list[tuple]:
    from openpyxl import load_workbook
    libro = load_workbook(io.BytesIO(bytes_xlsx))
    return list(libro.active.iter_rows(values_only=True))


print("un total simple (sin desglose) -- una sola fila con el numero")
filas = _leer(informes.generar_excel(
    "Clientes activos.", {"total": 4260, "interpretacion": "Clientes activos."}))
comprobar(filas[0][0] == "Clientes activos.", "la primera fila es la interpretacion")
comprobar(("Total", 4260) in filas, "el total aparece tal cual, sin recalcularlo")

print("\nun desglose -- una fila por categoria, ordenado de mayor a menor")
resultado = {"total": 7252, "desglose": {"activo": 4260, "suspendido": 811,
                                          "cancelado": 2093, "gratis": 88},
             "interpretacion": "Clientes, desglosado por estado."}
filas = _leer(informes.generar_excel(resultado["interpretacion"], resultado))
comprobar(("Categoria", "Cantidad") in filas, "hay encabezado de tabla")
comprobar(("activo", 4260) in filas, "el valor mas grande esta, tal cual")
comprobar(("gratis", 88) in filas, "el valor mas chico tambien, tal cual")
idx_activo = filas.index(("activo", 4260))
idx_gratis = filas.index(("gratis", 88))
comprobar(idx_activo < idx_gratis, "el desglose viene ordenado de mayor a menor")

print("\nsuma exacta: el archivo no inventa ni redondea nada")
suma_en_archivo = sum(v for _, v in filas if isinstance(v, int))
comprobar(suma_en_archivo == sum(resultado["desglose"].values()),
         "la suma de las filas del archivo coincide con el desglose original")


def _texto_pdf(bytes_pdf: bytes) -> str:
    from pypdf import PdfReader
    lector = PdfReader(io.BytesIO(bytes_pdf))
    return "\n".join(pagina.extract_text() for pagina in lector.pages)


print("\nPDF -- misma verificacion, mismo criterio: nada que el agregado no haya calculado")
pdf_total = informes.generar_pdf(
    "Clientes activos.", {"total": 4260, "interpretacion": "Clientes activos."})
comprobar(pdf_total[:5] == b"%PDF-", "el archivo es un PDF valido")
texto = _texto_pdf(pdf_total)
comprobar("Clientes activos." in texto, "la interpretacion esta en el PDF")
comprobar("4260" in texto, "el total aparece tal cual")

pdf_desglose = informes.generar_pdf(resultado["interpretacion"], resultado)
texto_desglose = _texto_pdf(pdf_desglose)
comprobar("activo" in texto_desglose and "4260" in texto_desglose,
         "el valor mas grande del desglose esta en el PDF")
comprobar("gratis" in texto_desglose and "88" in texto_desglose,
         "el valor mas chico tambien")
comprobar(texto_desglose.index("activo") < texto_desglose.index("gratis"),
         "el PDF tambien viene ordenado de mayor a menor")

if fallos:
    print(f"\n[FALLA] {len(fallos)} caso(s):")
    for f in fallos:
        print(f"  - {f}")
    sys.exit(1)

print("\n[OK] El archivo exportado refleja exactamente lo que el agregado calculo.")
